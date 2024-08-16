from django.shortcuts import render, get_object_or_404, redirect
from .models import Process
from .forms import KouteiEditForm, KouteiAddForm, MyModelForm, KouteiCommentForm, KouteiCopyForm
from django.views.generic import ListView
from django.db.models import Q, Max, Min, Sum, Avg, Count
import datetime
from django.db import connection
import openpyxl
from django.contrib import messages
import plotly.express as px
import pandas as pd
from pathlib import Path
from django.views.generic import TemplateView
from django_pandas.io import read_frame


# 検索
def top(request):
    # 初期値を設定
    # d = Process.objects.all().aggregate(Max('date'))
    # f = MyModelForm(initial={'date': d['date__max'], 'period': '昼勤'})
    d = datetime.datetime.today().strftime("%Y-%m-%d")
    f = MyModelForm(initial={'date': d, 'period': '昼勤'})
    return render(request, 'kouteikanri/top.html', {'form1': f})


# リダイレクト用
def redirect_a(request):
    if request.method == 'POST':
        date = request.POST['date']
        period = request.POST['period']
        return render(request, 'kouteikanri/redirect.html', {'date': date, 'period': period})


# 全ライン一覧
# 参考url: https://qiita.com/t-iguchi/items/827865481e82bb32ad04
def all_list(request):
    if request.method == 'POST':
        date = request.POST['date']
        period = request.POST['period']
        sql_text = (
                "SELECT line, period, "
                "sum(value) AS val_sum, "
                "sum(value * status) AS val_end_sum, "
                "sum(value) - sum(value * status) AS left_val, "
                "count(status) - sum(status) AS left_cnt, "
                "min(starty) AS starty_min, "
                "max(endy) AS endy_max, "
                "max(endj) AS endj_max, "
                "(sum(processy) + sum(changey)) - sum(processy * status) - sum(changey * status) AS left_time, "
                "(sum(processj * status) + sum(changej * status)) - (sum(processy * status) + sum(changey * status)) AS real_time, "
                "sum(value * status) * 100 / sum(value) AS progress "
                "FROM kouteikanri_process "
                "WHERE date='" + date +
                "' AND period='" + period +
                "' GROUP BY line, period "
                "ORDER BY period DESC, line;"
        )
        emp_list = exec_query(sql_text)
        return render(request, 'kouteikanri/all.html', {'emp_list': emp_list, 'date': date, 'period': period})
    else:
        # 通らないはず
        return redirect('kouteikanri:all', {'date': '2021-01-01', 'period': '昼勤'})


# セットチェック 全ライン一覧
#def set_all(request):
#    if request.method == 'POST':
#        date = request.POST['date']
#        period = request.POST['period']
#        sql_text = (
#                "SELECT line, period, "
#                "count(set) AS all_cnt, "
#                "sum(set) AS end_cnt, "
#                "count(set) - sum(set) AS left_cnt, "
#                "sum(set) * 100 / count(set) AS progress "
#                "FROM kouteikanri_process "
#                "WHERE date='" + date +
#                "' AND period='" + period +
#                "' AND hinban IS NOT NULL "
#                "GROUP BY line, period "
#                "ORDER BY period DESC, line;"
#        )
#        emp_list = exec_query(sql_text)
#        return render(request, 'kouteikanri/set_all.html', {
#            'emp_list': emp_list, 'date': date, 'period': period})


# cursor.descriptionでフィールド名を配列にセットして、resultsにフィールド名を付加
def exec_query(sql_txt):
    with connection.cursor() as c:
        c.execute(sql_txt)
        results = c.fetchall()
        columns = []
        for field in c.description:
            columns.append(field.name)
        values = []
        for result in results:
            value_dic = {}
            for index, field in enumerate(columns):
                value_dic[field] = result[index]
            values.append(value_dic)
        return values


# 工程一覧
class KouteiList(ListView):
    model = Process
    context_object_name = 'kouteis'
    template_name = 'kouteikanri/list.html'
    paginate_by = 50

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ln = self.kwargs['line']
        dt = self.kwargs['date']
        pr = self.kwargs['period']
        ql = Q(line__exact=ln)
        qd = Q(date__exact=dt)
        qp = Q(period__exact=pr)
        qs = Q(status__exact=1)
        # ライン名 ========================================================================
        ctx['linef'] = self.kwargs['line']
        # 製造日 ==========================================================================
        tstr = self.kwargs['date']
        tdata = datetime.datetime.strptime(tstr, '%Y-%m-%d')
        tdate = str(tdata.year) + '年' + str(tdata.month) + '月' + str(tdata.day) + '日'
        ctx['datef'] = tdate
        # 時間帯 ==========================================================================
        ctx['periodf'] = self.kwargs['period']
        # 終了予定 ========================================================================
        ctx['maxendy'] = endy_max(ln, dt, pr)
        # 終了数/総数 ========================================================================
        ctx['valsum'] = value_sum(ln, dt, pr)
        ctx['valend'] = value_end(ln, dt, pr)
        # 生産高 ===========================================================================
        sd = Process.objects.all().filter(ql & qd & qp).aggregate(Sum('seisand'))
        if sd['seisand__sum'] is None:
            sdstr = '0 円'
        else:
            sdstr = '{:,}'.format(round(sd['seisand__sum'] / 1000)) + ' 千円'
        ctx['sdstr'] = sdstr
        # 能率 =============================================================================
        stfavg = Process.objects.all().filter(
            ql & qd & qp & qs).aggregate(Avg('staff'))
        if stfavg['staff__avg'] is None:
            jin = 0
        else:
            jin = stfavg['staff__avg']
        sdEnd = Process.objects.all().filter(ql & qd & qp & qs).aggregate(Sum('seisand'))
        if sdEnd['seisand__sum'] is None:
            sdE = 0
        else:
            sdE = sdEnd['seisand__sum']
        minSt = Process.objects.all().filter(ql & qd & qp).aggregate(Min('startj'))
        maxEn = Process.objects.all().filter(ql & qd & qp).aggregate(Max('endj'))
        if minSt['startj__min'] is None:
            tm = 0
        else:
            Smin = minSt['startj__min']
            if maxEn['endj__max'] is not None:
                Emax = maxEn['endj__max']
                tm = get_stime(Smin, Emax) / 60
            else:
                tm = 0
        try:
            ctx['nouritsu'] = '{:,}'.format(round(sdE / tm / jin)) + ' 円'
        except ZeroDivisionError:
            ctx['nouritsu'] = '0 円'
        # 生産中の調査 ===================================================================
        cntst = Process.objects.all().filter(
            ql & qd & qp & Q(status__exact=0) & Q(startj__isnull=False)
        ).count()
        ctx['cntst'] = cntst
        # 切替平均 ======================================================================
        ctx['chavg'] = changej_avg(ln, dt, pr)
        # 終了予測 ======================================================================
        ctx['comptime'] = comp_time(ln, dt, pr)
        # 進捗 ==========================================================================
        if endy_max(ln, dt, pr) is None:
            progress = 0
        else:
            progress = get_stime(comp_time(ln, dt, pr), endy_max(ln, dt, pr))
        ctx['progress'] = progress
        ctx["new_date"] = datetime.datetime.now()
        return ctx

    def get_queryset(self, **kwargs):
        return Process.objects.order_by('startj', 'starty').filter(
            Q(line__exact=self.kwargs['line']) &
            Q(date__exact=self.kwargs['date']) &
            Q(period__exact=self.kwargs['period'])
        )

    @staticmethod
    def post(request):
        line = request.POST['line']
        date = request.POST['date']
        period = request.POST['period']
        return redirect('kouteikanri:list', line, date, period)


# セット状況
class SetList(ListView):
    model = Process
    context_object_name = 'kouteis'
    template_name = 'kouteikanri/set_list.html'
    paginate_by = 50

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ln = self.kwargs['line']
        dt = self.kwargs['date']
        pr = self.kwargs['period']
        # ライン名 ========================================================================
        ctx['linef'] = self.kwargs['line']
        # 製造日 ==========================================================================
        tstr = self.kwargs['date']
        tdata = datetime.datetime.strptime(tstr, '%Y-%m-%d')
        tdate = str(tdata.year) + '年' + str(tdata.month) + '月' + str(tdata.day) + '日'
        ctx['datef'] = tdate
        # 時間帯 ==========================================================================
        ctx['periodf'] = self.kwargs['period']
        # セット総数 =======================================================================
        ctx['setcnt'] = set_cnt(ln, dt, pr)
        # セット総数 =======================================================================
        ctx['setend'] = set_end(ln, dt, pr)
        # 進捗 ============================================================================
        ctx['progress'] = set_prog(ln, dt, pr)
        return ctx

    def get_queryset(self, **kwargs):
        return Process.objects.order_by('startj', 'starty').filter(
            Q(line__exact=self.kwargs['line']) &
            Q(date__exact=self.kwargs['date']) &
            Q(period__exact=self.kwargs['period']) &
            Q(hinban__gt=0))

    @staticmethod
    def post(request):
        line = request.POST['line']
        date = request.POST['date']
        period = request.POST['period']
        return redirect('kouteikanri:set_list', line, date, period)


# セット状況 (全ライン)
class SetListAll(ListView):
    model = Process
    context_object_name = 'kouteis'
    template_name = 'kouteikanri/set_list_all.html'
    paginate_by = 25
    d = datetime.datetime.today().strftime("%Y-%m-%d")
    form = MyModelForm(initial={'date': d, 'period': '昼勤'})

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        dt = self.kwargs['date']
        pr = self.kwargs['period']
        # 製造日 ==========================================================================
        tstr = self.kwargs['date']
        tdata = datetime.datetime.strptime(tstr, '%Y-%m-%d')
        tdate = str(tdata.year) + '年' + str(tdata.month) + '月' + str(tdata.day) + '日'
        ctx['datef'] = tdate
        # 時間帯 ==========================================================================
        ctx['periodf'] = self.kwargs['period']
        # セット総数 =======================================================================
        ctx['setcnt'] = set_cnt('*', dt, pr)
        # セット総数 =======================================================================
        ctx['setend'] = set_end('*', dt, pr)
        # 進捗 ============================================================================
        ctx['progress'] = set_prog('*', dt, pr)
        return ctx

    def get_queryset(self, **kwargs):
        return Process.objects.order_by('line', 'startj', 'starty').filter(
            Q(date__exact=self.kwargs['date']) &
            Q(period__exact=self.kwargs['period']) &
            Q(hinban__gt=0))


    @staticmethod
    def post(request):
        date = request.POST['date']
        period = request.POST['period']
        return render(request, 'kouteikanri:set_list_all', context={'date': date, 'period': period})


# 「総数」を計算する関数
def value_sum(line, date, period):
    koutei = Process.objects.filter(
        Q(line__exact=line) &
        Q(date__exact=date) &
        Q(period__exact=period)
    )
    if koutei.count() == 0:
        return 0
    else:
        return koutei.aggregate(Sum('value'))


# 「終了数」を計算する関数
def value_end(line, date, period):
    koutei = Process.objects.filter(
        Q(line__exact=line) &
        Q(date__exact=date) &
        Q(period__exact=period) &
        Q(status__exact=1)
    )
    if koutei.count() == 0:
        return 0
    else:
        return koutei.aggregate(Sum('value'))


# 「切替平均」を計算する関数
def changej_avg(line, date, period):
    koutei = Process.objects.filter(
        Q(line__exact=line) &
        Q(date__exact=date) &
        Q(period__exact=period)
    )
    if koutei.count() == 0:
        return 0
    else:
        return koutei.aggregate(Avg('changej'))


# 「終了予定」を取得する関数
def endy_max(line, date, period):
    koutei = Process.objects.filter(
        Q(line__exact=line) &
        Q(date__exact=date) &
        Q(period__exact=period)
    )
    if koutei.count() == 0:
        return None
    else:
        ym = koutei.aggregate(Max('endy'))
        if ym['endy__max'] is None:
            return None
        else:
            return ym['endy__max'].astimezone()


# 「終了実績」を取得する関数
def endj_max(line, date, period):
    koutei = Process.objects.filter(
        Q(line__exact=line) &
        Q(date__exact=date) &
        Q(period__exact=period)
    )
    if koutei.count() == 0:
        return None
    else:
        jm = koutei.aggregate(Max('endj'))
        if jm['endj__max'] is None:
            return None
        else:
            return jm['endj__max'].astimezone()


# 「残り生産時間」を計算する関数
def left_time(line, date, period):
    koutei = Process.objects.filter(
        Q(line__exact=line) &
        Q(date__exact=date) &
        Q(period__exact=period) &
        Q(status__exact=0)
    )
    sum_py = koutei.aggregate(Sum('processy'))
    if sum_py['processy__sum'] is None:
        py = 0
    else:
        py = sum_py['processy__sum']
    sum_cy = koutei.aggregate(Sum('changey'))
    if sum_cy['changey__sum'] is None:
        cy = 0
    else:
        cy = sum_cy['changey__sum']
    tt = py + cy
    return datetime.timedelta(minutes=tt)


# 「終了予測」を計算する関数
def comp_time(line, date, period):
    koutei = Process.objects.filter(
        Q(line__exact=line) &
        Q(date__exact=date) &
        Q(period__exact=period) &
        Q(status__exact=1)
    )
    if koutei.count() == 0:
        return None
    else:
        ct = endy_max(line, date, period)
        if ct is None:
            return None
        else:
            return endj_max(line, date, period) + left_time(line, date, period)


# 「セット総数」を計算する関数
def set_cnt(line, date, period):
    if line == '*':
        koutei = Process.objects.filter(
            Q(date__exact=date) &
            Q(period__exact=period) &
            Q(hinban__gt=0)
        )
    else:
        koutei = Process.objects.filter(
            Q(line__exact=line) &
            Q(date__exact=date) &
            Q(period__exact=period) &
            Q(hinban__gt=0)
        )
    if koutei.count() == 0:
        return 0
    else:
        cn = koutei.aggregate(Count('set'))
        return cn['set__count']


# 「セット完了」を計算する関数
def set_end(line, date, period):
    if line == '*':
        koutei = Process.objects.filter(
            Q(date__exact=date) &
            Q(period__exact=period) &
            Q(set__exact=1) &
            Q(hinban__gt=0)
        )
    else:
        koutei = Process.objects.filter(
            Q(line__exact=line) &
            Q(date__exact=date) &
            Q(period__exact=period) &
            Q(set__exact=1) &
            Q(hinban__gt=0)
        )
    if koutei.count() == 0:
        return 0
    else:
        sm = koutei.aggregate(Sum('set'))
        return sm['set__sum']


# 「セット進捗率」
def set_prog(line, date, period):
    if line == '*':
        koutei = Process.objects.filter(
            Q(date__exact=date) &
            Q(period__exact=period) &
            Q(hinban__gt=0)
        )
    else:
        koutei = Process.objects.filter(
            Q(line__exact=line) &
            Q(date__exact=date) &
            Q(period__exact=period) &
            Q(hinban__gt=0)
        )
    if koutei.count() == 0:
        return 0
    else:
        sm = koutei.aggregate(Sum('set'))
        cn = koutei.aggregate(Count('set'))
        if sm['set__sum'] is None:
            return 0
        else:
            if cn['set__count'] is None or cn['set__count'] == 0:
                return 0
            else:
                return round(sm['set__sum'] / cn['set__count'] * 100, 1)


# 新規 or 編集
def edit(request, id=None):
    # 編集
    if id:
        koutei = get_object_or_404(Process, pk=id)
        form = KouteiEditForm(request.POST, instance=koutei)
        template = 'kouteikanri/edit.html'
        # スライス枚数を変数に格納
        if koutei.value is not None and koutei.value != 0:
            if koutei.slicev is not None and koutei.slicev != 0:
                mai = koutei.slicev / koutei.value
            else:
                mai = 0
        else:
            mai = 0
    # 新規
    else:
        koutei = Process()
        form = KouteiAddForm(request.POST, instance=koutei)
        template = 'kouteikanri/add.html'
        mai = 0
    # POST
    if request.method == 'POST':
        # バリデーションチェック
        if form.is_valid():
            koutei = form.save(commit=False)
            koutei.line = request.POST['line']
            koutei.date = request.POST['date']
            koutei.period = request.POST['period']
            # 数量に応じて生産高と生産数hを再計算
            if koutei.value is not None:
                if koutei.price is not None:
                    koutei.seisand = koutei.value * koutei.price
                if koutei.seisanh is not None and koutei.seisanh != 0:
                    koutei.processy = round(koutei.value / koutei.seisanh * 60)
                if mai != 0:
                    koutei.slicev = koutei.value * mai
            # 終了時間に応じて実際時間を計算しstatusを更新
            if koutei.endj is not None:
                if koutei.startj is not None:
                    if koutei.name == '予備':
                        koutei.processj = koutei.processy
                    else:
                        koutei.processj = get_stime(koutei.startj, koutei.endj)
                koutei.status = 1
            else:
                koutei.processj = None
                koutei.status = 0
            # =============================================================================
            # ライン, 時間帯の変更があった場合と予備追加の場合のみfkeyを新規に付番する
            # =============================================================================
            old_line = koutei.tracker.previous('line')
            old_period = koutei.tracker.previous('period')
            if koutei.fkey is None or koutei.line is not old_line \
                    or koutei.period is not old_period:
                koutei_f = Process.objects.filter(
                    Q(line__exact=koutei.line) &
                    Q(date__exact=koutei.date) &
                    Q(period__exact=koutei.period) &
                    Q(bin__exact=koutei.bin) &
                    Q(name=koutei.name)
                )
                nm = koutei_f.count() + 1
                koutei.fkey = nm
            # 保存
            koutei.save()
            # =============================================================================
            # 例えば fkey の末尾 1 の工程のラインを変更したとき末尾 2 以降の工程があった場合には
            # 付番しなおす必要がある
            # =============================================================================
            update_list = []
            i = 1
            koutei_old = Process.objects.filter(
                Q(line__exact=old_line) &
                Q(date__exact=koutei.date) &
                Q(period__exact=old_period) &
                Q(bin__exact=koutei.bin) &
                Q(name=koutei.name)
            )
            if koutei_old.count() > 0:
                for koutei in koutei_old:
                    koutei.fkey = i
                    i = i + 1
                    update_list.append(koutei)
                # データベースを更新
                Process.objects.bulk_update(update_list, fields=["fkey"])
            # =============================================================================
            if 'next' in request.GET:
                return redirect(request.GET['next'])
        else:
            print("validation error: id=" + str(id))
    # GET
    else:
        if id:
            form = KouteiEditForm(instance=koutei)
            template = 'kouteikanri/edit.html'
        else:
            form = KouteiAddForm(instance=koutei)
            template = 'kouteikanri/add.html'
        if 'next' in request.GET:
            return redirect(request.GET['next'])
    return render(request, template, dict(form=form, id=id))


# 再生産
def copy(request, id=None):
    # POST
    koutei = Process()
    form = KouteiCopyForm(request.POST, instance=koutei)
    if request.method == 'POST':
        koutei_f = Process.objects.filter(
            Q(line__exact=request.POST['line']) &
            Q(date__exact=request.POST['date']) &
            Q(period__exact=request.POST['period']) &
            Q(name__exact=request.POST['name'])
        )
        nm = koutei_f.count() + 1
        # バリデーションチェック
        if form.is_valid():
            koutei = form.save(commit=False)
            koutei.line = request.POST['line']
            koutei.date = request.POST['date']
            koutei.period = request.POST['period']
            # 数量に応じて生産高と生産数hを再計算
            if koutei.value is not None and koutei.value != 0:
                if koutei.price is not None:
                    koutei.seisand = koutei.value * koutei.price
                if koutei.seisanh is not None and koutei.seisanh != 0:
                    koutei.processy = round(koutei.value / koutei.seisanh * 60)
            # 開始状態にする
            koutei.startj = datetime.datetime.now().astimezone()
            koutei.status = 0
            koutei.fkey = nm
            # 保存
            koutei.save()
            if 'next' in request.GET:
                return redirect(request.GET['next'])
        return redirect('kouteikanri:list', request.POST['line'], request.POST['date'], request.POST['period'])
    # GET
    else:
        koutei = get_object_or_404(Process, pk=id)
        initial_dict = dict(
            line=koutei.line, date=koutei.date, period=koutei.period,
            bin=koutei.bin, hinban=koutei.hinban, price=koutei.price,
            kubun='再生産', name=koutei.name, seisanh=koutei.seisanh,
            value=0, seisand=0, conveyor=koutei.conveyor, staff=koutei.staff,
            panmm=koutei.panmm, slicev=0, slicep=koutei.slicep,
            changey=koutei.changey, processy=0, status=0
        )
        form = KouteiCopyForm(initial=initial_dict)
        if 'next' in request.GET:
            return redirect(request.GET['next'])
    return render(request, 'kouteikanri/edit.html', dict(form=form))


# 終了解除
def end_none(request, id=id):
    koutei = get_object_or_404(Process, pk=id)
    if request.method == 'POST':
        koutei.endj = None
        # 終了を解除、開始はそのまま
        koutei.processj = None
        koutei.status = 0
        koutei.save()
    return redirect('kouteikanri:list', koutei.line, koutei.date, koutei.period)


# 削除
def delete(request, id):
    koutei = get_object_or_404(Process, pk=id)
    koutei.delete()
    # return redirect('kouteikanri:list', koutei.line, koutei.date, koutei.period)
    return redirect(request.META.get('HTTP_REFERER', '/', ))


# 開始 or 終了
def start_or_end(request, id=id):
    koutei = get_object_or_404(Process, pk=id)
    # 開始の処理
    ql = Q(line__exact=koutei.line)
    qd = Q(date__exact=koutei.date)
    qp = Q(period__exact=koutei.period)
    if koutei.startj is None:
        nw = datetime.datetime.now().astimezone()
        # 切替時間を更新
        if koutei.value is not None:
            maxendj = Process.objects.all().filter(ql & qd & qp).aggregate(Max('endj'))
            if maxendj['endj__max'] is None:
                koutei.changej = 0
            else:
                dt = maxendj['endj__max'].astimezone()
                koutei.changej = get_stime(dt, nw)
        # 開始時間を更新
        koutei.startj = nw
        koutei.save()
    else:
        if koutei.endj is None:
            koutei.endj = datetime.datetime.now().astimezone()
            if koutei.name == '予備':
                koutei.processj = koutei.processy
            else:
                koutei.processj = get_stime(koutei.startj, koutei.endj)
        koutei.status = 1
        koutei.save()
    # return redirect('kouteikanri:list', koutei.line, koutei.date, koutei.period)
    return redirect(request.META.get('HTTP_REFERER', '/', ))


# 生産中をキャンセル
def start_cancel(request, **kwargs):
    # POST
    if request.method == 'POST':
        update_list = []
        line = request.POST['line']
        date = request.POST['date']
        dt = datetime.datetime.strptime(date, '%Y年%m月%d日')
        if dt is None:
            d = datetime.datetime.now().strftime("%Y-%m-%d")
        else:
            d = dt.strftime("%Y-%m-%d")
        period = request.POST['period']
        kouteis = Process.objects.all().filter(
            Q(line__exact=line) &
            Q(date__exact=d) &
            Q(period__exact=period) &
            Q(status__exact=0) &
            Q(endj__isnull=True)
        )
        if kouteis.count() > 0:
            for koutei in kouteis:
                koutei.startj = None
                koutei.changej = None
                koutei.status = 0
                update_list.append(koutei)
            # 生産中のデータを一括更新
            Process.objects.bulk_update(
                update_list, fields=["changej", "startj", "status"])
        # return redirect('kouteikanri:list', line, d, period)
        return redirect(request.META.get('HTTP_REFERER', '/', ))


# セット完了
def set_comp(request, id=id):
    koutei = get_object_or_404(Process, pk=id)
    # 完了の処理
    if koutei.set == 0:
        nw = datetime.datetime.now().astimezone()
        koutei.set = 1
        koutei.setj = nw
    # 解除の処理
    else:
        koutei.set = 0
        koutei.setj = None
    koutei.save()
    # return redirect('kouteikanri:set_list', koutei.line, koutei.date, koutei.period)
    return redirect(request.META.get('HTTP_REFERER', '/', ))


# すべての実績をリセット
def reset_all(request, **kwargs):
    # POST
    if request.method == 'POST':
        update_list = []
        line = request.POST['line']
        date = request.POST['date']
        period = request.POST['period']
        kouteis = Process.objects.all().filter(
            Q(line__exact=line) &
            Q(date__exact=date) &
            Q(period__exact=period)
        )
        if kouteis.count() > 0:
            for koutei in kouteis:
                koutei.startj = None
                koutei.endj = None
                koutei.changej = None
                koutei.processj = None
                koutei.status = 0
                update_list.append(koutei)
            # 工程のデータを一括更新
            Process.objects.bulk_update(
                update_list, fields=["startj", "endj", "changej", "processj", "status"])
        # return redirect('kouteikanri:list', line, date, period)
        return redirect(request.META.get('HTTP_REFERER', '/', ))


# 所要時間計算
def get_stime(start_time, end_time):
    if start_time is None:
        return 0
    elif end_time is None:
        return 0
    else:
        td = end_time - start_time
        return round((td.days * 1440) + (td.seconds / 60))


# 備考
def comment(request, id):
    koutei = get_object_or_404(Process, pk=id)
    form = KouteiCommentForm(instance=koutei)
    if 'next' in request.GET:
        return redirect(request.GET['next'])
    else:
        return render(request, 'kouteikanri/comment.html', {'form': form})


# アップロード
def upload(request):
    if request.method == 'POST' and request.FILES['excel']:
        excel = request.FILES['excel']
        if not excel.name[-4:] == 'xlsx':
            print(excel.name[-4:])
            messages.error(request, "Excelファイル(*.xlsx)を選択してください")
            return render(request, 'kouteikanri/upload.html')
        # Excelの読み込み
        wb = openpyxl.load_workbook(excel)
        for ws in wb.worksheets:
            # シートの最終行を取得
            sheet_max_row = ws.max_row
            # ライン名, 時間帯, 製造日を抽出
            line = ws.title
            date_str = ws.cell(1, 1).value
            if not date_str[:3] == "製造日":
                messages.error(request, "シート" + ws.title + "は工程表ではありません")
                continue
            else:
                messages.info(request, "ライン「" + ws.title + "」の工程をアップロード")
            period = date_str[-2:]
            date_trm = str(date_str[4:date_str.index('(')])
            date_y = str(date_trm[0:4])
            date_m = str(date_trm[date_trm.index('年') + 1:date_trm.index('月')])
            date_d = str(date_trm[date_trm.index('月') + 1:date_trm.index('日')])
            date_ymd = datetime.date(int(date_y), int(date_m), int(date_d))
            name_list = []
            update_list = []
            # 開始実績がないデータに削除フラグを立てる
            koutei_delete = Process.objects.filter(
                Q(line__exact=line) &
                Q(date__exact=date_ymd) &
                Q(period__exact=period) &
                Q(startj__isnull=True) &
                Q(status__exact=0)
            )
            if koutei_delete.count() > 0:
                for koutei in koutei_delete:
                    koutei.status = -1
                    update_list.append(koutei)
                # データベースを更新
                Process.objects.bulk_update(update_list, fields=["status"])
            # 行をループ
            for j in range(6, sheet_max_row + 1):
                binn = ws.cell(row=j, column=2).value
                name = ws.cell(row=j, column=6).value
                # 切替予定がNoneなら0に置換する
                if ws.cell(row=j, column=15).value is None:
                    chy = 0
                else:
                    chy = ws.cell(row=j, column=15).value
                starty = ws.cell(row=j, column=17).value
                if name == '合計':
                    messages.warning(request, "　")
                else:
                    # fkey
                    name_list.append(name + str(binn))
                    fkey = name_list.count(name + str(binn))
                    # 開始予定と同じ終了予定のデータがあれば工程が連続していると判断し合算処理
                    koutei_gassan = Process.objects.filter(
                        Q(line__exact=line) &
                        Q(date__exact=date_ymd) &
                        Q(period__exact=period) &
                        Q(bin__exact=binn) &
                        Q(name__exact=name) &
                        Q(hinban__isnull=False) &
                        Q(endy__exact=starty)
                    )
                    if koutei_gassan.count() == 1 and fkey > 1:
                        messages.warning(request, "　　合算：" + name)
                        koutei = Process.objects.get(
                            line=line,
                            date=date_ymd,
                            period=period,
                            bin=binn,
                            name=name,
                            hinban__isnull=False,
                            endy=starty
                        )
                        koutei.kubun = '確定'
                        koutei.seisanh = ws.cell(row=j, column=7).value
                        if koutei.value:
                            if ws.cell(row=j, column=8).value:
                                koutei.value = koutei.value + ws.cell(row=j, column=8).value
                                koutei.seisand = koutei.seisand + ws.cell(row=j, column=9).value
                        if koutei.slicev:
                            if ws.cell(row=j, column=13).value:
                                koutei.slicev = koutei.slicev + ws.cell(row=j, column=13).value
                        koutei.processy = koutei.processy + ws.cell(row=j, column=16).value
                        koutei.endy = ws.cell(row=j, column=18).value
                        koutei.save()
                    else:
                        koutei_fkey = Process.objects.filter(
                            Q(line__exact=line) &
                            Q(date__exact=date_ymd) &
                            Q(period__exact=period) &
                            Q(bin__exact=binn) &
                            Q(name__exact=name) &
                            Q(fkey__exact=fkey)
                        )
                        # 同じfkeyのデータがDBにない場合
                        if koutei_fkey.count() == 0:
                            # 新規追加
                            messages.warning(request, "　　追加：" + name)
                            Process.objects.create(
                                line=line,
                                period=period,
                                date=date_ymd,
                                bin=binn,
                                hinban=ws.cell(row=j, column=3).value,
                                price=ws.cell(row=j, column=4).value,
                                kubun=ws.cell(row=j, column=5).value,
                                name=name,
                                seisanh=ws.cell(row=j, column=7).value,
                                value=ws.cell(row=j, column=8).value,
                                seisand=ws.cell(row=j, column=9).value,
                                conveyor=ws.cell(row=j, column=10).value,
                                staff=ws.cell(row=j, column=11).value,
                                panmm=ws.cell(row=j, column=12).value,
                                slicev=ws.cell(row=j, column=13).value,
                                slicep=ws.cell(row=j, column=14).value,
                                changey=chy,
                                processy=ws.cell(row=j, column=16).value,
                                starty=ws.cell(row=j, column=17).value,
                                endy=ws.cell(row=j, column=18).value,
                                status=0,
                                fkey=fkey
                            )
                        # 同じfkeyのデータがDBにある場合
                        elif koutei_fkey.count() == 1:
                            koutei = Process.objects.get(
                                line=line,
                                date=date_ymd,
                                period=period,
                                bin=binn,
                                name=name, fkey=fkey
                            )
                            # 予測過剰は更新しない
                            messages.warning(request, "　　更新：" + name)
                            if koutei.kubun == '予測' and koutei.endj is not None \
                                    and koutei.value > ws.cell(row=j, column=8).value:
                                messages.error(request, "　　　予測過剰")
                            else:
                                koutei.kubun = ws.cell(row=j, column=5).value
                                koutei.seisanh = ws.cell(row=j, column=7).value
                                koutei.value = ws.cell(row=j, column=8).value
                                koutei.seisand = ws.cell(row=j, column=9).value
                            # 常に更新する
                            koutei.price = ws.cell(row=j, column=4).value
                            koutei.conveyor = ws.cell(row=j, column=10).value
                            koutei.staff = ws.cell(row=j, column=11).value
                            koutei.panmm = ws.cell(row=j, column=12).value
                            koutei.slicev = ws.cell(row=j, column=13).value
                            koutei.slicep = ws.cell(row=j, column=14).value
                            koutei.changey = chy
                            koutei.processy = ws.cell(row=j, column=16).value
                            koutei.starty = ws.cell(row=j, column=17).value
                            koutei.endy = ws.cell(row=j, column=18).value
                            if koutei.status == -1:
                                koutei.status = 0
                            koutei.save()
                        else:
                            messages.error(
                                request,
                                "　" + name + "には" + str(koutei_fkey.count()) +
                                "件の検索キーが存在しています"
                            )
            # 削除フラグが残っているデータを削除
            koutei_delete = Process.objects.filter(Q(status__exact=-1))
            if koutei_delete.count() > 0:
                for koutei in koutei_delete:
                    messages.error(request, "　削除：" + koutei.name)
                    koutei.delete()
        messages.success(request, "ファイルのアップロードが終了しました")
    return render(request, 'kouteikanri/upload.html')


def line_charts(date, period):
    # CSV読み込み　※デバッグ時のみ使用
    # p = Path("C:/Users/Administrator/Desktop")
    # df = pd.read_csv(p.joinpath("file.csv"), header=0,
    #                  index_col=0, engine="python", encoding="utf_8_sig")

    koutei = Process.objects.order_by('line').filter(
        Q(date__exact=date) &
        Q(period__exact=period)
    )
    df = read_frame(koutei)

    # datetune列の内容をDatetime型かつaware(UTC)で、"datetime_utc"列に入れる
    df["start_utc"] = pd.to_datetime(df["starty"], utc=True)
    df["end_utc"] = pd.to_datetime(df["endy"], utc=True)

    # JSTに変換したデータを"datetime_jst"列に入れる
    df["start_jst"] = df["start_utc"].dt.tz_convert('Asia/Tokyo')
    df["end_jst"] = df["end_utc"].dt.tz_convert('Asia/Tokyo')

    fig = px.timeline(
        df, x_start="start_jst", x_end="end_jst", y="line", text="name",
        color="status",
        color_continuous_scale=["aliceblue", "palegreen"],
        labels={'start_jst': '開始', 'end_jst': '終了', 'line': 'ライン', 'name': '製品名',
                'status': '状態'},
        # height=1440,
    )
    fig.update_traces(
        width=0.95,
        textposition='inside',
        textfont=dict(color='black'),
    )
    fig.update_layout(
        uniformtext_mode=False,
        uniformtext_minsize=10,
        plot_bgcolor='blanchedalmond',
        clickmode='none',
    )
    fig.update_xaxes(
        title=dict(text='', font=dict(color='grey')),
        gridcolor='black', gridwidth=1,
    )
    fig.update_yaxes(
        title=dict(text='', font=dict(color='grey')),
        gridcolor='white', gridwidth=1,
        categoryorder='array',
        categoryarray=df['line'][::-1],
    )

    return fig.to_html(include_plotlyjs=False)


class LineChartsView(TemplateView):
    model = Process
    context_object_name = 'kouteis'
    template_name = 'kouteikanri/plot.html'
    d = datetime.datetime.today().strftime("%Y-%m-%d")
    form = MyModelForm(initial={'date': d, 'period': '昼勤'})

    def get_context_data(self, **kwargs):
        context = super(LineChartsView, self).get_context_data(**kwargs)
        dt = self.kwargs['date']
        pr = self.kwargs['period']
        context["plot"] = line_charts(dt, pr)
        # 製造日 ==========================================================================
        tdata = datetime.datetime.strptime(dt, '%Y-%m-%d')
        tdate = str(tdata.year) + '年' + str(tdata.month) + '月' + str(tdata.day) + '日'
        context['datef'] = tdate
        return context

    def get_queryset(self, **kwargs):
        return Process.objects.order_by('line', 'starty').filter(
            Q(date__exact=self.kwargs['date']) &
            Q(period__exact=self.kwargs['period']) &
            Q(hinban__gt=0)
        )

    @staticmethod
    def post(request):
        date = request.POST['date']
        period = request.POST['period']
        return render(request, 'kouteikanri:plot', context={'date': date, 'period': period})
